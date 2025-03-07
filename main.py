from typing import List, Dict

from z3 import *
import openpyxl

from vacation_date_to_week_index import vacation_date_to_week_index

# Not too dangerous to make global
W = 52

def range_fellows_assigned_fully(o, x, R, fellow_start, fellow_end):
    # Each NCC fellow has exactly one rotation per week, because we are responsible for their scheduleo.
    for f in range(fellow_start, fellow_end):
        for w in range(W):
            o.add(AtLeast(*[x[f, w, r] for r in R], 1))


def everyone_one_rotation_per_week(o, x, R, fellow_start, fellow_end):
    # Each other fellow has at most one rotation per week, since we are only assigning their NCC time.
    for f in range(fellow_start, fellow_end):
        for w in range(W):
            o.add(AtMost(*[x[f, w, r] for r in R], 1))


def ncc_shifts_covered_swing_deficit(o, x, N, deficit):
    # There is one fellow on Swing and at least one fellow on NCC1, NCC2 per week.
    # We can be more specific if this gets nuts with overassignment.
    for w in range(W):
        o.add(Sum([If(x[f, w, "NCC1"], 1, 0) for f in range(N)]) >= 1)
        o.add(Sum([If(x[f, w, "NCC2"], 1, 0) for f in range(N)]) >= 1)
        o.add(Sum([If(x[f, w, "NCC1"], 1, 0) for f in range(N)]) <= 2)
        o.add(Sum([If(x[f, w, "NCC2"], 1, 0) for f in range(N)]) <= 2)
        # At most one extra fellow on at once.
        o.add(Sum([If(Or(x[f, w, "NCC1"], x[f, w, "NCC2"]), 1, 0) for f in range(N)]) <= 3)

        # LE because inadequacy
        o.add(Sum([If(x[f, w, "Swing"], 1, 0) for f in range(N)]) <= 1)

    # Ah, we might not actually have enough swing. Let's say at most 8 weeks are
    # unassigned.
    o.add(Sum([
        Sum([If(x[f, w, "Swing"], 1, 0) for f in range(N)])
        for w in range(W)]) >= W-deficit)

def ncc_stroke_oversight(o, x, fellow_start, fellow_end):
    # IDEALLY every week either NCC1 or NCC2 is neurocrit or stroke.
    for w in range(W):
        o.add(
            Sum([
                If(
                    Or(x[f, w, "NCC1"], x[f, w, "NCC2"]),
                    1,
                    0
                ) for f in range(fellow_start, fellow_end)
            ]) >= 1
        )

def maximum_consecutive_icu_shifts(o, x, fellow_start, fellow_end, MAX_CONSEC):

    for f in range(fellow_start, fellow_end):
        for w in range(W - MAX_CONSEC):
            # for r in ["NCC1", "NCC2", "Swing", "SICU", "MICU"]:
            #     o.add(Sum([If(x[f, w + i, r], 1, 0) for i in range(MAX_CONSEC + 1)]) <= MAX_CONSEC)
            o.add(Sum([If(
                Or(*[x[f, w + i, r] for r in ["NCC1", "NCC2", "Swing", "SICU", "MICU"]]),
                1,
                0) for i in range(MAX_CONSEC + 1)]) <= MAX_CONSEC)

def jr_first_month_micu(o, x, fellow_start, fellow_end):
    # jr fellows first month is MICU
    for f in range(fellow_start, fellow_end):
        for w in range(4):
            o.add(x[f, w, "MICU"])

def jr_ncc_before_19(o, x, fellow_start, fellow_end):
    # jr fellows have a block of NCC before week 19
    for f in range(fellow_start, fellow_end):
        o.add(
            Sum([If(Or(x[f, w, "NCC1"], x[f, w, "NCC2"]), 1, 0) for w in range(4, 19)]) >= 1
        )

def ccm_total_service(o, x, fellow_start, fellow_end):
    # ccm fellows have precisely one month of NCC, of which one week is swing
    # TODO: follow 'block' boundaries
    for f in range(fellow_start, fellow_end):
        # oh, and it's consecutive.
        # actually to do this, it's just as easy to do blocks

        # Total for the year
        o.add(Sum([If(Or(x[f, w, "NCC1"], x[f, w, "NCC2"]), 1, 0) for w in range(52)]) == 3)
        o.add(Sum([If(x[f, w, "Swing"], 1, 0) for w in range(52)]) == 1)

        # Consecutivity
        for w in range(0, W, 4):
            # If the first week of the block is NCC-ish, the rest must be.
            # Oh no that is insane logic. Either all or none of the block is NCC-ish.
            o.add(
                If(
                    Sum([If(Or(x[f, w_, "NCC1"], x[f, w_, "NCC2"], x[f, w_, "Swing"]), 1, 0) for w_ in range(w, w + 4)]) > 0,
                    Sum([If(Or(x[f, w_, "NCC1"], x[f, w_, "NCC2"], x[f, w_, "Swing"]), 1, 0) for w_ in
                         range(w, w + 4)]),
                    4) == 4 # if the whole block isn't NCCish, this doesn't fail condition
            )
            o.add(
                If(
                    Sum([If(Or(x[f, w_, "NCC1"], x[f, w_, "NCC2"], x[f, w_, "Swing"]), 1, 0) for w_ in
                         range(w, w + 4)]) > 0,
                    Sum([If(x[f, w_, "Swing"], 1, 0) for w_ in
                         range(w, w + 4)]),
                    1) == 1  # if the whole block isn't NCCish, this doesn't fail condition
            )

def total_shift_service(o, x, f, shift, n):
    o.add(Sum([If(x[f, w, shift], 1, 0) for w in range(W)]) >= n)

def total_nicu_service(o, x, f, n):
    o.add(Sum([If(Or(x[f, w, "NCC1"], x[f, w, "NCC2"]), 1, 0) for w in range(W)]) >= n)

def stroke_total_service(o, x, fellow_start, fellow_end):
    for f in range(fellow_start, fellow_end):
        total_shift_service(o, x, f, "Swing", 2)
        total_nicu_service(o, x, f, 6)

def ncc_jr_total_service(o, x, fellow_start, fellow_end):
    for f in range(fellow_start, fellow_end):
        total_shift_service(o, x, f, "MICU", 20)
        total_shift_service(o, x, f, "Anaesthesia", 4)
        total_shift_service(o, x, f, "Elec", 9)
        total_shift_service(o, x, f, "Vac", 3)
        total_shift_service(o, x, f, "NS", 0)
        total_shift_service(o, x, f, "SICU", 4)
        total_shift_service(o, x, f, "Vasc/Clin", 0)
        total_shift_service(o, x, f, "Swing", 3) # not sure how this was 6 TODO
        total_nicu_service(o, x, f, 9)

def ncc_jr_total_service(o, x, fellow_start, fellow_end):
    for f in range(fellow_start, fellow_end):
        total_shift_service(o, x, f, "MICU", 20)
        total_shift_service(o, x, f, "Anaesthesia", 4)
        total_shift_service(o, x, f, "Elec", 9)
        total_shift_service(o, x, f, "Vac", 3)
        total_shift_service(o, x, f, "NS", 0)
        total_shift_service(o, x, f, "SICU", 4)
        total_shift_service(o, x, f, "Vasc/Clin", 0)
        total_shift_service(o, x, f, "Swing", 3)  # not sure how this was 6 TODO
        total_nicu_service(o, x, f, 9)

def ncc_sr_total_service(o, x, fellow_start, fellow_end):
    for f in range(fellow_start, fellow_end):
        total_shift_service(o, x, f, "MICU", 8)
        total_shift_service(o, x, f, "Anaesthesia", 0)
        total_shift_service(o, x, f, "Elec", 10)
        total_shift_service(o, x, f, "Vac", 3)
        total_shift_service(o, x, f, "NS", 7)
        total_shift_service(o, x, f, "SICU", 0)
        total_shift_service(o, x, f, "Vasc/Clin", 4)
        total_shift_service(o, x, f, "Swing", 6)
        total_nicu_service(o, x, f, 14)
        pass

# def Min(arg):
#     return If(
#         len(arg) == 0,
#         arg[0],
#         If(
#             arg[0] > arg[1],
#             Min(arg[1:]),
#             Min(arg[0]+arg[2:])))

def jr_fellows_n_ncc_before_swing(o, x, fellow_start, fellow_end, n):
    # jr fellows have 4x NCC before their first swing
    # Or(x[f, w, "NCC1"], x[f, w, "NCC2"]),

    for f in range(fellow_start, fellow_end):
        o.add(
            Sum([
                Product([
                    # Number of NCC shifts before week w.
                    Sum([
                        If(
                            Or(x[f, w_, "NCC1"], x[f, w_, "NCC2"]),
                            1,
                            0
                        )
                    for w_ in range(w)]),
                    # Zero if there is a swing shift before week w, or if week w itself is not a swing shift.
                    If(
                        And(
                            Sum([
                                If(
                                    x[f, w_, "Swing"],
                                    1,
                                    0
                                ) for w_ in range(w)
                            ]) == 0,
                            x[f,w,"Swing"],
                        ),
                        1,
                        0
                    )
                ])
            for w in range(W)]) >= n
        )

def shift_blocked(o, x, shift, fellow_min, fellow_max, GRANULARITY):
    # junior fellows have 4 sicu, and it should follow a block.
    # TODO: for now we are requiring 4 block

    for f in range(fellow_min, fellow_max):
        # Consecutivity
        for w in range(0, W, GRANULARITY):
            # Either all or none of the block is SICU.
            o.add(
                Or(
                    Sum([
                        If(x[f, w_, shift], 1, 0) for w_ in range(w, w + GRANULARITY)
                    ]) == GRANULARITY,
                    Sum([
                        If(x[f, w_, shift], 1, 0) for w_ in range(w, w + GRANULARITY)
                    ]) == 0,
                )
            )

def sicu_blocked(o, x, fellow_start, fellow_end):
    # junior fellows have 4 sicu, and it should follow a block.
    # TODO: for now we are requiring 4 block
    shift_blocked(o, x, "SICU", fellow_start, fellow_end, GRANULARITY = 4)

def micu_blocked(o, x, fellow_start, fellow_end):
    # jr and sr fellows have lots of micu, and it should follow a block.
    # TODO: for now we are requiring 2 block
    shift_blocked(o, x, "MICU", fellow_start, fellow_end, GRANULARITY = 2)


def anaesthesia_blocked(o, x, fellow_start, fellow_end):
    # jr and sr fellows have lots of micu, and it should follow a block.
    # TODO: for now we are requiring 4 block
    shift_blocked(o, x, "Anaesthesia", fellow_start, fellow_end, GRANULARITY = 4)

def vasc_blocked(o, x, fellow_start, fellow_end):
    # jr and sr fellows have lots of micu, and it should follow a block.
    # TODO: for now we are requiring 4 block
    shift_blocked(o, x, "Vasc/Clin", fellow_start, fellow_end, GRANULARITY = 4)

def ns_blocked(o, x, fellow_start, fellow_end):
    # jr and sr fellows have lots of micu, and it should follow a block.
    # TODO: for now we are requiring 4 block
    GRANULARITY = 4
    shift = "NS"

    for f in range(fellow_start, fellow_end):
        # Consecutivity
        for w in range(0, W, GRANULARITY):
            # Either all or none of the block is SICU.
            o.add(
                Or(
                    Sum([
                        If(x[f, w_, shift], 1, 0) for w_ in range(w, w + GRANULARITY)
                    ]) == GRANULARITY,
                    Sum([
                        If(x[f, w_, shift], 1, 0) for w_ in range(w, w + 3)
                    ]) == 3,
                    Sum([
                        If(x[f, w_, shift], 1, 0) for w_ in range(w, w + GRANULARITY)
                    ]) == 0,
                )
            )


def ncc_blocked(o, x, fellow_start, fellow_end):
    # TODO: for now we are requiring 2 block
    GRANULARITY = 2
    for f in range(fellow_start, fellow_end):
        # Consecutivity
        for w in range(0, W, GRANULARITY):
            o.add(
                Or(
                    Sum([
                        If(Or(x[f, w_, "NCC1"], x[f, w_, "NCC2"], x[f, w_, "Swing"]), 1, 0) for w_ in range(w, w + GRANULARITY)
                    ]) == GRANULARITY,
                    Sum([
                        If(Or(x[f, w_, "NCC1"], x[f, w_, "NCC2"], x[f, w_, "Swing"]), 1, 0) for w_ in range(w, w + GRANULARITY)
                    ]) == 0,
                )
            )

def vacation_requests(o, x, fellows, fellow_week_pairs, n_vac):
    # prash wants weeks 1, 7, and 36
    # (figure out a way to express this TODO)
    for f_, w_ in fellow_week_pairs.items():
        f = fellows.index(f_)
        for w in w_[:n_vac]:
            o.add(
                x[f, w, "Vac"]
            )
        for w in w_[n_vac:]:
            o.add(
                x[f, w, "Elec"]
            )

def fourth_block_two_micu_fellows(o, x, fellow_start, fellow_end):
    # from the fellows between start and end, ensure MICU is double-staffed for every week from 12-15
    for w in range(12,16):
        o.add(
            Sum([If(x[f, w, "MICU"], 1, 0) for f in range(fellow_start, fellow_end)]) == 2,
        )

def comparable_amounts_each_half_year(o, x, fellow_start, fellow_end):
    # I don't want any shift to be massively frontloaded or backloaded.
    # no one's year should end with 8 NCC, 2 Elec, 8 NCC, 2 Elec, 8 NCC
    for f in range(fellow_start, fellow_end):
        o.add(
            Abs(
                # Diff(
                    Sum([
                        If(x[f, w, "MICU"], 1, 0) for w in range(0, W // 2)
                    ]) -
                    Sum([
                        If(x[f, w, "MICU"], 1, 0) for w in range(W // 2, W)
                    ])
                # )
            ) <= 4
        )

        o.add(
            Abs(
                # Diff(
                    Sum([
                        If(Or(x[f, w, "NCC1"], x[f, w, "NCC2"], x[f, w, "Swing"]), 1, 0) for w in range(0, W // 2)
                    ]) -
                    Sum([
                        If(Or(x[f, w, "NCC1"], x[f, w, "NCC2"], x[f, w, "Swing"]), 1, 0) for w in range(W // 2, W)
                    ])
                # )
            ) <= 4
        )
    pass

def optimize_schedule(
    jr_fellows: List[str],
    sr_fellows: List[str],
    stroke_fellows: List[str],
    CCM_fellows: List[str],
    R: List[str],
    fellow_week_pairs: Dict[str, List[int]],
):
    fellows = jr_fellows + sr_fellows + stroke_fellows + CCM_fellows

    num_NCC_jr_fellows = len(jr_fellows)
    num_NCC_sr_fellows = len(sr_fellows)
    num_stroke_fellows = len(stroke_fellows)
    num_CCM_fellows = len(CCM_fellows)
    N = num_NCC_jr_fellows + num_NCC_sr_fellows + num_stroke_fellows + num_CCM_fellows  # Number of fellows (example)


    # A 3D boolean variable: x[f, w, r] is True if fellow f is assigned to rotation r in week w
    x = {
        (f, w, r): Bool(f"x_{f}_{w}_{r}") for f in range(N) for w in range(W) for r in R
    }

    # s = Solver()
    o = Optimize()


    # assign NCC fellows fully
    range_fellows_assigned_fully(o, x, R, fellow_start=0, fellow_end=num_NCC_jr_fellows+num_NCC_sr_fellows)
    everyone_one_rotation_per_week(o, x, R, fellow_start=0, fellow_end=N)
    ncc_shifts_covered_swing_deficit(o, x, N,8)
    maximum_consecutive_icu_shifts(o, x, fellow_start=0, fellow_end=N, MAX_CONSEC=8)
    jr_first_month_micu(o, x, fellow_start=0, fellow_end=num_NCC_jr_fellows)
    jr_ncc_before_19(o, x, fellow_start=0, fellow_end=num_NCC_jr_fellows)
    jr_fellows_n_ncc_before_swing(o, x, fellow_start=0, fellow_end=num_NCC_jr_fellows, n=4)
    fourth_block_two_micu_fellows(o, x, fellow_start=0, fellow_end=num_NCC_jr_fellows+num_NCC_sr_fellows)

    sicu_blocked(o,x, fellow_start=0, fellow_end=num_NCC_jr_fellows)
    micu_blocked(o,x, fellow_start=0, fellow_end=num_NCC_jr_fellows+num_NCC_sr_fellows)
    anaesthesia_blocked(o,x, fellow_start=0, fellow_end=num_NCC_jr_fellows)
    vasc_blocked(o,x, fellow_start=num_NCC_jr_fellows, fellow_end=num_NCC_jr_fellows+num_NCC_sr_fellows)
    ns_blocked(o,x, fellow_start=num_NCC_jr_fellows, fellow_end=num_NCC_jr_fellows+num_NCC_sr_fellows)
    ncc_blocked(o,x, fellow_start=0, fellow_end=N)
    ncc_stroke_oversight(o,x, fellow_start = 0, fellow_end=num_NCC_jr_fellows + num_NCC_sr_fellows + num_stroke_fellows)
    vacation_requests(o,x, fellows, fellow_week_pairs, n_vac=3)
    comparable_amounts_each_half_year(o, x, fellow_start=0, fellow_end=num_NCC_jr_fellows + num_NCC_sr_fellows)


    """
    sum over each fellow.
    """
    ccm_total_service(o,x, fellow_start=num_NCC_jr_fellows + num_NCC_sr_fellows + num_stroke_fellows, fellow_end=N)
    stroke_total_service(o,x,fellow_start=num_NCC_jr_fellows + num_NCC_sr_fellows, fellow_end=num_NCC_jr_fellows + num_NCC_sr_fellows + num_stroke_fellows)
    ncc_jr_total_service(o,x, fellow_start=0, fellow_end=num_NCC_jr_fellows)
    ncc_sr_total_service(o,x, fellow_start=num_NCC_jr_fellows, fellow_end=num_NCC_jr_fellows+num_NCC_sr_fellows)

    print(o.check())

    m = o.model()

    shifts_for_fellows = {
        fellow: ["" for w in range(W)] for fellow in fellows
    }

    fellows_for_shifts = {
        'NCC1': [[] for w in range(W)],
        'NCC2': [[] for w in range(W)],
        'Extra': [[] for w in range(W)],
        'Swing': [[] for w in range(W)],
    }

    for d in m.decls():
        if m[d]:
            _, f, w, s = d.name().split('_')
            shifts_for_fellows[fellows[int(f)]][int(w)] = s
            if s in fellows_for_shifts:
                fellows_for_shifts[s][int(w)].append(fellows[int(f)])

            # print ("%s = %s" % (d.name(), m[d]))

    # figure out who's extra and who isn't
    for s, v in fellows_for_shifts.items():
        for ii, its in enumerate(v):
            if type(its) is list and len(its) == 0:
                # print(s, 0)
                fellows_for_shifts[s][ii] = ""
                # fellows_for_shifts['Extra'][ii] = ""
                continue
            elif type(its) is list and len(its) == 1:
                # print(s, 1)
                fellows_for_shifts[s][ii] = its[0]
                # fellows_for_shifts['Extra'][ii] = ""
                continue
            elif type(its) is list and len(its) == 2:
                # print(len(its), its)
                # extra hierarchy: CCM is more extra than stroke is more extra than NCC natives.
                fellows_for_shifts['Extra'][ii] = max(its)
                fellows_for_shifts[s][ii] = min(its)
            else:
                # print(s, ii, its)
                continue
    # print(fellows_for_shifts['Extra'])

    return shifts_for_fellows, fellows_for_shifts

if __name__ == "__main__":

    jr_fellows = ["NCC Raya", "NCC Joseph"]
    sr_fellows = ["NCC David", "NCC Prash"]
    stroke_fellows = ["Stroke Arthur", "Stroke Betty", "Stroke Charles", "Stroke Deirdre"]
    CCM_fellows = ["CCM Ariana", "CCM Bert", "CCM Chloe", "CCM Dennis", "CCM Edwina", "CCM Frank", "CCM George",
                   "CCM Helen", "CCM Iago", "CCM Jake", "CCM Kyle", "CCM Liana", "CCM Mary", "CCM Ning", "CCM Oyo"]

    R = ["NCC1", "NCC2", "Swing", "SICU", "MICU", "Elec", "Vac", "NS", "Vasc/Clin", "Anaesthesia"]  # Example rotations
    fellows = jr_fellows + sr_fellows + stroke_fellows + CCM_fellows

    fellow_week_pairs = {
        "NCC Prash": [1, 7, 36,
                       vacation_date_to_week_index((2026, 3, 30)),
                       vacation_date_to_week_index((2026, 4, 6)),
                       vacation_date_to_week_index((2026, 4, 13)),
                       vacation_date_to_week_index((2026, 4, 20)),
                       ],
        "NCC David": [5, 6, 28],
        "NCC Raya": [21, 37],
        "NCC Joseph": [
            vacation_date_to_week_index((2025, 12, 25)),
            vacation_date_to_week_index((2026, 3, 9)),
            vacation_date_to_week_index((2025, 10, 20)),
            vacation_date_to_week_index((2026, 5, 18)),
            vacation_date_to_week_index((2025, 9, 16)) # elective for ABPN
        ],
    }

    shifts_for_fellows, fellows_for_shifts = optimize_schedule(
        jr_fellows, sr_fellows, stroke_fellows, CCM_fellows, R, fellow_week_pairs,
    )

    std_output = False

    if std_output:
        # print out total schedules for the NCC fellows
        print('    ,' + ','.join(fellows[:len(jr_fellows) + len(sr_fellows)]))
        for w in range(W):
            print(f'{w: 4d},', end='')
            for i in range(len(jr_fellows) + len(sr_fellows)):
                print(shifts_for_fellows[fellows[i]][w], end=",")
            print()

        # print out total schedules for the NCC fellowship shifts
        print('    ,' + ','.join(['NCC1', 'NCC2', 'Extra', 'Swing']))
        for w in range(W):
            print(f'{w: 4d},', end='')
            for s in ['NCC1', 'NCC2', 'Extra', 'Swing']:
                print('+'.join(fellows_for_shifts[s][w]), end=",")
            print('\n')
    else:

        nsYYMMDD = openpyxl.styles.NamedStyle(name="cd1", number_format="YYYY-MM-DD")

        # thick_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thick'),
        #                      right=openpyxl.styles.borders.Side(style='thick'),
        #                      top=openpyxl.styles.borders.Side(style='thick'),
        #                      bottom=openpyxl.styles.borders.Side(style='thick'))
        thick_side = openpyxl.styles.borders.Side(style='thick')
        def compose_borders(cell, top=None, bottom=None, left=None, right=None):
            # print(top, bottom, left, right)
            cell.border = openpyxl.styles.borders.Border(top=top, bottom=bottom, left=left, right=right)

        fellow_fill_dict = {}
        shift_fill_dict = {}
        micu_color = "B4C6E7"
        ncc_color = "C6E0B4"
        stroke_color = "F8CBAD"
        swing_color = "A9D08E"
        vasc_color = "FCE4D6"
        sicu_color = "FFE699"
        ns_color = "FFF3CC"
        anaesthesia_color = "F8CBAD"
        for f in jr_fellows:
            fellow_fill_dict[f] = openpyxl.styles.PatternFill(
                start_color=ncc_color, end_color=ncc_color, fill_type='solid')
        for f in sr_fellows: fellow_fill_dict[f] = openpyxl.styles.PatternFill(
                start_color=ncc_color, end_color=ncc_color, fill_type='solid')
        for f in stroke_fellows: fellow_fill_dict[f] = openpyxl.styles.PatternFill(
                start_color=stroke_color, end_color=stroke_color, fill_type='solid')
        for f in CCM_fellows: fellow_fill_dict[f] = openpyxl.styles.PatternFill(
                start_color=micu_color, end_color=micu_color, fill_type='solid')

        shift_fill_dict['MICU'] = openpyxl.styles.PatternFill(
            start_color=micu_color, end_color=micu_color, fill_type='solid')
        shift_fill_dict['SICU'] = openpyxl.styles.PatternFill(
            start_color=sicu_color, end_color=sicu_color, fill_type='solid')
        shift_fill_dict['NCC1'] = openpyxl.styles.PatternFill(
            start_color=ncc_color, end_color=ncc_color, fill_type='solid')
        shift_fill_dict['NCC2'] = openpyxl.styles.PatternFill(
            start_color=ncc_color, end_color=ncc_color, fill_type='solid')
        shift_fill_dict['Swing'] = openpyxl.styles.PatternFill(
            start_color=swing_color, end_color=swing_color, fill_type='solid')
        shift_fill_dict['Vasc/Clin'] = openpyxl.styles.PatternFill(
            start_color=vasc_color, end_color=vasc_color, fill_type='solid')
        shift_fill_dict['NS'] = openpyxl.styles.PatternFill(
            start_color=ns_color, end_color=ns_color, fill_type='solid')
        shift_fill_dict['Anaesthesia'] = openpyxl.styles.PatternFill(
            start_color=anaesthesia_color, end_color=anaesthesia_color, fill_type='solid')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Per-Fellow Schedule"

        n_fellows_to_output = len(jr_fellows) + len(sr_fellows)
        column_letters = "BCDEFGHIJKLMNOPQRSTUVWXYZ"[:n_fellows_to_output]
        for cl, fellow in zip(column_letters, fellows[:n_fellows_to_output]):
            ws[f'{cl}1'] = fellow

        ws[f'A1'] = 'Week'
        ws[f'A2'] = '2025-06-30' # TODO: coordinate with vacation date file constant
        ws[f'A2'].style = nsYYMMDD
        for windex in range(3, W+2):
            ws[f'A{windex}'] = f'=A{windex-1}+7'
            ws[f'A{windex}'].style = nsYYMMDD
        for windex in range(2, W + 2):
            w = windex - 2
            for cl, fellow in zip(column_letters, fellows[:n_fellows_to_output]):
                # print(f'formatting {cl}{windex}')

                if (windex-2)%4 == 0:
                    if cl == 'B': compose_borders(ws[f'{cl}{windex}'], top=thick_side, left=thick_side)
                    elif cl == column_letters[n_fellows_to_output-1]: compose_borders(ws[f'{cl}{windex}'], top=thick_side, right=thick_side)
                    else: compose_borders(ws[f'{cl}{windex}'], top=thick_side)
                elif (windex-1)%4 == 0:
                    if cl == 'B': compose_borders(ws[f'{cl}{windex}'], bottom=thick_side, left=thick_side)
                    elif cl == column_letters[n_fellows_to_output-1]: compose_borders(ws[f'{cl}{windex}'], bottom=thick_side, right=thick_side)
                    else:
                        compose_borders(ws[f'{cl}{windex}'], bottom=thick_side)
                else:
                    if cl == 'B':
                        compose_borders(ws[f'{cl}{windex}'], left=thick_side)
                    elif cl == column_letters[n_fellows_to_output - 1]:
                        compose_borders(ws[f'{cl}{windex}'], right=thick_side)

                ws[f'{cl}{windex}'] = shifts_for_fellows[fellow][w]
                if shifts_for_fellows[fellow][w] in shift_fill_dict:
                    ws[f'{cl}{windex}'].fill = shift_fill_dict[shifts_for_fellows[fellow][w]]

        wb.create_sheet("NCC Shift Schedule")
        ws = wb["NCC Shift Schedule"]

        shifts = ["NCC1", "NCC2", "Extra", "Swing"]
        column_letters = "BCDEFGHIJKLMNOPQRSTUVWXYZ"[:len(shifts)]
        for cl, shift in zip(column_letters, shifts):
            ws[f'{cl}1'] = shift

        ws[f'A1'] = 'Week'
        ws[f'A2'] = '2025-06-30' # TODO: coordinate with vacation date file constant
        ws[f'A2'].style = nsYYMMDD
        for windex in range(3, W+2):
            ws[f'A{windex}'] = f'=A{windex-1}+7'
            ws[f'A{windex}'].style = nsYYMMDD

        for windex in range(2, W + 2):
            w = windex - 2
            for cl, s in zip(column_letters, shifts):


                if (windex-2)%4 == 0:
                    if cl == 'B': compose_borders(ws[f'{cl}{windex}'], top=thick_side, left=thick_side)
                    elif cl == column_letters[len(shifts)-1]: compose_borders(ws[f'{cl}{windex}'], top=thick_side, right=thick_side)
                    else: compose_borders(ws[f'{cl}{windex}'], top=thick_side)
                elif (windex-1)%4 == 0:
                    if cl == 'B': compose_borders(ws[f'{cl}{windex}'], bottom=thick_side, left=thick_side)
                    elif cl == column_letters[len(shifts)-1]: compose_borders(ws[f'{cl}{windex}'], bottom=thick_side, right=thick_side)
                    else:
                        compose_borders(ws[f'{cl}{windex}'], bottom=thick_side)
                else:
                    if cl == 'B':
                        compose_borders(ws[f'{cl}{windex}'], left=thick_side)
                    elif cl == column_letters[n_fellows_to_output - 1]:
                        compose_borders(ws[f'{cl}{windex}'], right=thick_side)

                ws[f'{cl}{windex}'] = fellows_for_shifts[s][w]
                if fellows_for_shifts[s][w] in fellow_fill_dict:
                    ws[f'{cl}{windex}'].fill = fellow_fill_dict[fellows_for_shifts[s][w]]

        wb.save(filename="optimized_schedule.xlsx")